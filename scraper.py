import asyncio
from playwright.async_api import async_playwright
import pandas as pd
import random
import re
import os
from datetime import datetime
import unicodedata
from openpyxl.styles import Alignment

def clean_text(text):
    if not text: return text
    # Normalize unicode characters
    text = unicodedata.normalize('NFKC', text)
    # Replace common windows-1252/unicode quotation mark issues that excel struggles with
    text = text.replace('"', '"').replace('"', '"').replace('"', '"')
    text = text.replace(''', "'").replace(''', "'")
    text = text.replace('...', '...')
    text = text.replace('—', '-').replace('–', '-')
    # Remove zero width spaces and other invisible formatting characters
    text = re.sub(r'[\u200b\u200e\u200f\u202a-\u202e\ufeff]', '', text)
    return text.strip()

async def extract_list_data(page):
    books = []
    print("Extracting data from Bestsellers list...")
    
    # Wait for the grid of items to load
    try:
        await page.wait_for_selector('div[id^="p13n-asin-index-"]', timeout=10000)
    except:
        print("Timeout waiting for grid. Page content might be different or CAPTCHA presented.")
        return []

    items = await page.query_selector_all('div[id^="p13n-asin-index-"]')
    print(f"Found {len(items)} items on the page.")

    for item in items:
        book_data = {
            'Rank': None,
            'Title': None,
            'Author': None,
            'Rating': None,
            'Reviews': None,
            'Price': None,
            'URL': None
        }

        try:
            rank_el = await item.query_selector('.zg-bdg-text')
            if not rank_el:
                rank_el = await item.query_selector('.zg-badge-text')
            if rank_el:
                book_data['Rank'] = clean_text(await rank_el.inner_text())
            if not book_data['Rank']:
                # Regex fallback on full item text
                full_text = await item.inner_text()
                match = re.search(r'#(\d+)', full_text)
                if match:
                    book_data['Rank'] = f"#{match.group(1)}"
        except: pass

        try:
            title_el = await item.query_selector('._cDEzb_p13n-sc-css-line-clamp-1_1Fn1y')
            if title_el:
                book_data['Title'] = clean_text(await title_el.inner_text())
        except: pass

        try:
            author_el = await item.query_selector('.a-row.a-size-small .a-link-child')
            if not author_el:
                author_el = await item.query_selector('.a-row.a-size-small .a-color-base')
            if author_el:
                book_data['Author'] = clean_text(await author_el.inner_text())
        except: pass

        try:
            rating_el = await item.query_selector('.a-icon-alt')
            if rating_el:
                rating_text = await rating_el.inner_text()
                match = re.search(r'([\d.]+)\s*out of', rating_text)
                if match:
                    book_data['Rating'] = float(match.group(1))
        except: pass

        try:
            # The most robust way is finding the product-reviews link and extracting aria-label or inner text
            reviews_link = await item.query_selector('a[href*="product-reviews"]')
            if reviews_link:
                # Option 1: extract from aria-label (e.g., "4.7 out of 5 stars, 5,195 ratings")
                aria_label = await reviews_link.get_attribute('aria-label')
                if aria_label:
                    match = re.search(r',\s*([\d,]+)\s*ratings', aria_label)
                    if match:
                        book_data['Reviews'] = int(match.group(1).replace(',', '').strip())
                
                # Option 2: Extract from the span inside if aria-label failed
                if not book_data['Reviews']:
                    reviews_span = await reviews_link.query_selector('.a-size-small')
                    if reviews_span:
                        reviews_text = await reviews_span.inner_text()
                        cleaned_reviews = clean_text(reviews_text).replace(',', '').strip()
                        if cleaned_reviews.isdigit():
                            book_data['Reviews'] = int(cleaned_reviews)
        except: pass

        try:
            price_el = await item.query_selector('._cDEzb_p13n-sc-price_3mJ9Z')
            if not price_el:
                price_el = await item.query_selector('.p13n-sc-price')
            if price_el:
                price_text = await price_el.inner_text()
                match = re.search(r'[\d.]+', price_text)
                if match:
                    book_data['Price'] = float(match.group(0))
        except: pass

        try:
            link_el = await item.query_selector('a.a-link-normal')
            if link_el:
                href = await link_el.get_attribute('href')
                if href:
                    if href.startswith('http'):
                        book_data['URL'] = href
                    else:
                        book_data['URL'] = "https://www.amazon.com" + href
        except: pass
        
        if book_data['Title'] and book_data['URL']:
            books.append(book_data)

    return books

async def extract_book_details(page, url):
    details = {
        'Description': None,
        'Publisher': None,
        'Publication Date': None
    }
    
    print(f"Visiting product page: {url}")
    try:
        await page.goto(url, timeout=30000, wait_until='domcontentloaded')
        await asyncio.sleep(random.uniform(2.0, 4.0))
        
        if await page.query_selector('#captchacharacters'):
            print("CAPTCHA encountered on product page!")
            return details

        try:
            desc_el = await page.query_selector('#bookDescription_feature_div')
            if not desc_el:
                desc_el = await page.query_selector('#productDescription')
            if desc_el:
                desc_text = await desc_el.inner_text()
                details['Description'] = clean_text(desc_text.replace('\n', ' '))
        except: pass

        try:
            bullets = await page.query_selector_all('#detailBullets_feature_div li')
            for bullet in bullets:
                text = await bullet.inner_text()
                if 'Publisher' in text:
                    cleaned_text = text.replace('Publisher', '').replace(':', '').strip()
                    match = re.search(r'(.*?)\s*\((.*?)\)', cleaned_text)
                    if match:
                        details['Publisher'] = clean_text(match.group(1))
                        pub_date_str = clean_text(match.group(2))
                        # We handle date parsing below
                        details['Publication Date'] = pub_date_str
                    else:
                        details['Publisher'] = clean_text(cleaned_text)
                
                if 'Publication date' in text or 'Publication Date' in text:
                    pub_date_str = clean_text(text.split(':')[-1].strip())
                    details['Publication Date'] = pub_date_str

            # Amazon's newer product layout check (Carousel/Grid formatting)
            if not details['Publisher']:
                pub_grid_el = await page.query_selector('#rpi-attribute-book_details-publisher .rpi-attribute-value')
                if pub_grid_el:
                    details['Publisher'] = clean_text(await pub_grid_el.inner_text())
            if not details['Publication Date']:
                date_grid_el = await page.query_selector('#rpi-attribute-book_details-publication_date .rpi-attribute-value')
                if date_grid_el:
                    details['Publication Date'] = clean_text(await date_grid_el.inner_text())
                    
            # Parse Date cleanly if we have one
            if details['Publication Date']:
                try:
                    parsed_date = datetime.strptime(details['Publication Date'], '%B %d, %Y')
                    details['Publication Date'] = parsed_date.strftime('%Y-%m-%d')
                except:
                    pass

        except: pass

    except Exception as e:
        print(f"Error visiting {url}: {e}")

    return details

async def run_scraper(target_url, max_items=20):
    """
    Runs the Playwright scraper on the given target_url and returns the path to the saved Excel file.
    """
    output_filename = 'scraped_data.xlsx'
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={'width': 1280, 'height': 800}
        )
        page = await context.new_page()

        print(f"Navigating to bestsellers page: {target_url}")
        await page.goto(target_url, timeout=30000)
        await asyncio.sleep(random.uniform(2.0, 5.0))
        
        if await page.query_selector('#captchacharacters'):
             print("CAPTCHA encountered on main page! Terminating.")
             await browser.close()
             return None

        for _ in range(3):
            await page.keyboard.press('PageDown')
            await asyncio.sleep(1.0)

        books = await extract_list_data(page)

        print(f"Successfully extracted {len(books)} books from the list.")
        
        if not books:
            await browser.close()
            return None

        books_to_process = books[:max_items]
        print(f"Fetching details for the top {len(books_to_process)} books...")

        all_book_data = []

        for book in books_to_process:
            details = await extract_book_details(page, book['URL'])
            combined = {**book, **details}
            all_book_data.append(combined)

        print("Finished scraping individual pages.")

        df = pd.DataFrame(all_book_data)
        cols = ['Rank', 'Title', 'Author', 'Rating', 'Reviews', 'Price', 'URL', 'Description', 'Publisher', 'Publication Date']
        final_cols = [c for c in cols if c in df.columns]
        df = df[final_cols]
        
        output_path = os.path.join(os.getcwd(), output_filename)
        
        # Use pandas ExcelWriter to customize formatting so Excel displays it cleanly
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Bestsellers')
            worksheet = writer.sheets['Bestsellers']
            
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            default_alignment = Alignment(vertical='top')
            
            for idx, col in enumerate(df.columns):
                # Calculate required column width
                max_len = max(df[col].fillna("").map(lambda x: len(str(x))).max() if not df[col].empty else 0, len(str(col)))
                # Don't let columns get insanely wide (cap at 60 characters)
                max_len = min(max_len, 60) 
                
                col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                worksheet.column_dimensions[col_letter].width = max_len + 2
                
                # Apply word wrap to long text columns and top-align everything
                is_long_text = col in ['Title', 'Description', 'URL', 'Publisher', 'Author']
                for row in worksheet.iter_rows(min_col=idx+1, max_col=idx+1, min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = wrap_alignment if is_long_text else default_alignment

        print(f"Data saved to {output_path}")

        await browser.close()
        
        return output_filename

if __name__ == "__main__":
    test_url = "https://www.amazon.com/Best-Sellers-Kindle-Store-Paranormal-Romance/zgbs/digital-text/154606011"
    asyncio.run(run_scraper(test_url, max_items=5))
